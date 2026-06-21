import os
import sys
import json
import glob

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.label_merge import merge_label

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "tea_flower_datasets")
LABELS_DIR = os.path.join(DATA_DIR, "labels")
TRAIN_LABELS_DIR = os.path.join(LABELS_DIR, "train")
VAL_LABELS_DIR = os.path.join(LABELS_DIR, "val")

FEATURES = ["花型", "花瓣皱褶", "花瓣外侧主色的颜色", "花瓣着色类型"]


def collect_labels(label_dirs):
    original_counts = {feature: {} for feature in FEATURES}
    merged_to_orig = {feature: {} for feature in FEATURES}
    total_samples = 0

    for label_dir in label_dirs:
        if not os.path.exists(label_dir):
            continue
        json_files = glob.glob(os.path.join(label_dir, "*.json"))
        for json_file in json_files:
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    label = json.load(f)
                total_samples += 1
                for feature in FEATURES:
                    value = label.get(feature, "")
                    if value and value.strip():
                        orig = value.strip()
                        if orig not in original_counts[feature]:
                            original_counts[feature][orig] = 0
                        original_counts[feature][orig] += 1

                        merged = merge_label(feature, orig)
                        if merged:
                            if merged not in merged_to_orig[feature]:
                                merged_to_orig[feature][merged] = []
                            if orig not in merged_to_orig[feature][merged]:
                                merged_to_orig[feature][merged].append(orig)
            except Exception as e:
                print(f"Error loading {json_file}: {e}")

    return original_counts, merged_to_orig, total_samples


def build_merged_structure(original_counts, merged_to_orig):
    result = {}
    for feature in FEATURES:
        feature_data = {}
        merged_labels = sorted(
            merged_to_orig[feature].keys(),
            key=lambda m: sum(original_counts[feature].get(o, 0) for o in merged_to_orig[feature][m]),
            reverse=True
        )
        for merged in merged_labels:
            orig_labels = sorted(
                merged_to_orig[feature][merged],
                key=lambda o: original_counts[feature].get(o, 0),
                reverse=True
            )
            total = sum(original_counts[feature].get(o, 0) for o in orig_labels)
            feature_data[merged] = {
                "total": total,
                "original": {o: original_counts[feature].get(o, 0) for o in orig_labels}
            }
        result[feature] = feature_data
    return result


def print_console_result(structure, total_samples):
    print()
    print("=" * 70)
    print("  合并后标签统计（含原始标签列表 + 数量）")
    print("=" * 70)

    for feature in FEATURES:
        feature_data = structure[feature]
        total_merged = len(feature_data)
        total_orig = sum(len(v["original"]) for v in feature_data.values())
        total_count = sum(v["total"] for v in feature_data.values())

        print()
        print(f"  【{feature}】")
        print(f"  合并后 {total_merged} 个标签  (原始 {total_orig} 个标签 / {total_count} 个样本)")
        print("  " + "-" * 60)
        print(f'  "{feature}": {{')

        for merged_label, data in feature_data.items():
            orig_items = list(data["original"].items())
            orig_str = ", ".join(f'"{o}"' for o, _ in orig_items)
            print(f'      "{merged_label}": [{orig_str}],   // {data["total"]} 个样本')

        print("  },")

    print()
    print(f"  总样本数: {total_samples}")
    print("=" * 70)


def export_excel(structure, total_samples, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    merged_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    merged_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws = wb.create_sheet(title="合并标签总览")
    ws.append(["特征", "合并后标签", "样本数", "占比", "包含原始标签数", "原始标签列表"])
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    current_row = 2
    for feature in FEATURES:
        feature_data = structure[feature]
        feature_total = sum(v["total"] for v in feature_data.values())

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"  【{feature}】  共 {len(feature_data)} 个合并标签 / {feature_total} 个样本"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = left_align
        cell.border = thin_border
        current_row += 1

        for idx, (merged_label, data) in enumerate(feature_data.items(), 1):
            orig_list = list(data["original"].keys())
            orig_str = "、".join(orig_list)
            ratio = f'{data["total"]/feature_total*100:.2f}%' if feature_total > 0 else "0%"

            ws.cell(row=current_row, column=1, value=idx).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align

            c = ws.cell(row=current_row, column=2, value=merged_label)
            c.fill = merged_fill
            c.font = merged_font
            c.border = thin_border

            ws.cell(row=current_row, column=3, value=data["total"]).border = thin_border
            ws.cell(row=current_row, column=3).alignment = center_align

            ws.cell(row=current_row, column=4, value=ratio).border = thin_border
            ws.cell(row=current_row, column=4).alignment = center_align

            ws.cell(row=current_row, column=5, value=len(orig_list)).border = thin_border
            ws.cell(row=current_row, column=5).alignment = center_align

            ws.cell(row=current_row, column=6, value=orig_str).border = thin_border
            ws.cell(row=current_row, column=6).alignment = left_align

            current_row += 1

        current_row += 1

    for col_letter, width in zip(["A", "B", "C", "D", "E", "F"], [6, 20, 10, 10, 16, 60]):
        ws.column_dimensions[col_letter].width = width

    for feature in FEATURES:
        sheet_name = feature[:28] + "详情" if len(feature) > 28 else f"{feature}详情"
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["合并后标签", "序号", "原始标签", "原始标签样本数", "合并后总样本数"])
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        feature_data = structure[feature]
        current_row = 2
        for merged_label, data in feature_data.items():
            orig_items = list(data["original"].items())
            start_row = current_row

            for idx, (orig_label, orig_count) in enumerate(orig_items, 1):
                ws.cell(row=current_row, column=2, value=idx).border = thin_border
                ws.cell(row=current_row, column=3, value=orig_label).border = thin_border
                ws.cell(row=current_row, column=4, value=orig_count).border = thin_border
                current_row += 1

            end_row = current_row - 1
            if start_row != end_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)

            c = ws.cell(row=start_row, column=1, value=merged_label)
            c.fill = merged_fill
            c.font = merged_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

            c2 = ws.cell(row=start_row, column=5, value=data["total"])
            c2.fill = merged_fill
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = thin_border

        for col_letter, width in zip(["A", "B", "C", "D", "E"], [22, 8, 35, 16, 16]):
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(output_path)
        print()
        print(f"  Excel 已保存至: {output_path}")
        print()
        print("  工作表列表:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"    {i}. {name}")
    except PermissionError:
        print()
        print(f"权限错误: 无法写入 {output_path}")
        print("  请先关闭 Excel 中打开的 all_labels.xlsx 文件，然后重新运行脚本。")
        sys.exit(1)


def main():
    label_dirs = [TRAIN_LABELS_DIR, VAL_LABELS_DIR]
    original_counts, merged_to_orig, total_samples = collect_labels(label_dirs)
    structure = build_merged_structure(original_counts, merged_to_orig)

    print_console_result(structure, total_samples)

    output_path = os.path.join(BASE_DIR, "all_labels.xlsx")
    export_excel(structure, total_samples, output_path)


if __name__ == "__main__":
    main()
